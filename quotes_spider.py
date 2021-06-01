
import scrapy
import xlrd
from openpyxl import load_workbook
import pandas as pd

class QuotesSpider(scrapy.Spider):
    name = "test1"
    start_urls = [
        'https://www.dallascad.org',

    ]

    def parse(self, response):
        web = 'https://www.dallascad.org/AcctDetailRes.aspx?ID='
        web1 = 'https://www.dallasact.com/act_webdev/dallas/showdetail2.jsp?can='
        web2 = 'https://www.dallasact.com/act_webdev/dallas/reports/taxbyyear.jsp?can='
        web3 = 'https://www.dallasact.com/act_webdev/dallas/reports/paymentinfo.jsp?can='

        loc = '/Users/ARK/Desktop/test1.xlsx'

        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(0)
        sheet.cell_value(0, 0)

        col_counter = 2
        while col_counter in range(sheet.nrows):
            print("Value of COunter is ", col_counter)
            Id = sheet.cell_value(col_counter, 0)
            if Id is not None:
                web_name = web + Id
                web_name1 = web1+Id
                web_name2 = web2+Id
                web_name3 = web3+Id
                # print(web_name)
                # print("Value of counter is ", counter)
                yield response.follow(web_name, callback=self.parse1)
                yield response.follow(web_name1, callback=self.parse2)
                yield response.follow(web_name2, callback=self.parse3)
                yield response.follow(web_name3, callback=self.parse4)
            col_counter += 1
        # for counter in range(sheet.nrows):
        #     print("Value of COunter is ",counter)
        #     Id = sheet.cell_value(counter, 0)
        #     if Id is not None:
        #         web_name = web + Id
        #         # print(web_name)
        #         # print("Value of counter is ", counter)
        #         yield response.follow(web_name, callback=self.parse1)

    def parse1(self, response):
        # Deleted Account" under column AF and leave all other cell of same
        # Start by opening the spreadsheet and selecting the main sheet
        workbook = load_workbook(filename="/Users/ARK/Desktop/output1.xlsx")
        sheet = workbook.active

        # # Write what you want into a specific cell
        # i = 1
        # while (i < 50):
        #     index = "B" + repr(i)
        #     print(index)
        #     sheet["B" + repr(i)] = "Reading"
        #     i += 1
        # # Save the spreadsheet
        # workbook.save(filename="/Users/ARK/Desktop/output.xlsx")

        Account = response.css('#lblPageTitle::text').get()
        counter =1

        if Account == "Deleted Account":
            print("Value of Account is ", Account)
            print("Nothing")
            AF = "Deleted Account"
        else:

            #     Owner
            #     response.xpath('//span[@id="lblOwner"]/following-sibling::text()').extract()[0]
            #     response.xpath('//span[@id="lblOwner"]/following-sibling::text()').extract()[1]
            #     response.xpath('//span[@id="lblOwner"]/following-sibling::text()').extract()[2].split(',')

            #                   Owner Detail
            # EW = response.xpath('//span[@id="lblOwner"]/following-sibling::text()').extract()[0]
            # value = response.xpath('//span[@id="lblOwner"]/following-sibling::text()').extract()[1]
            # for character in value:
            #     if character.isdigit():
            #         EZ = response.xpath('//span[@id="lblOwner"]/following-sibling::text()').extract()[1]
            #     else:
            #         EY = response.xpath('//span[@id="lblOwner"]/following-sibling::text()').extract()[1]
            #         FD = response.xpath('//span[@id="lblOwner"]/following-sibling::text()').extract()[2].split(',')


            Address1 = response.css('#LegalDesc1_lblLegal1::text').get()
            Address2 = response.css('#LegalDesc1_lblLegal2::text').get()
            Address3 = response.css('#LegalDesc1_lblLegal3::text').get()
            Address4 =response.css('#LegalDesc1_lblLegal4::text').get()
            Address5 = response.css('#LegalDesc1_lblLegal5::text').get()

            if(Address1==None):
                Up_Address1 = response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[0].strip()
            else:
                Up_Address1 = response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[0].strip() + response.css('#LegalDesc1_lblLegal1::text').get()

            if (Address2 == None):
                Up_Address2 = response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[1].strip()
            else:
                Up_Address2 = response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[1].strip() + response.css('#LegalDesc1_lblLegal2::text').get()


            if (Address3 == None):
                Up_Address3 = response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[2].strip()
            else:
                Up_Address3 =response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[2].strip() + response.css('#LegalDesc1_lblLegal3::text').get()


            if (Address4 == None):
                Up_Address4 = response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[3].strip()
            else:
                Up_Address4 = response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[3].strip() + response.css('#LegalDesc1_lblLegal4::text').get()

            if (Address5 == None):
                Up_Address5 = response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[4].strip()
            else:
                Up_Address5 = response.xpath('//table[@id="Table8"]/tr/th/text()').getall()[4].strip() + response.css('#LegalDesc1_lblLegal5::text').get().replace(' ','')

            IA = Up_Address1+','+Up_Address2+','+Up_Address3+','+Up_Address4+','+Up_Address5
            KS =  response.css('#LegalDesc1_lblSaleDate::text').get()
            KR = response.css('#LegalDesc1_lblSaleDate::text').get()[5:]

            # Multi-Owner (Current 2021)
            AZ = response.xpath('//table[@id="MultiOwner1_dgmultiOwner"]/tr[2]/td[2]/text()').extract_first()

            # Value
            O =  response.css('#ValueSummary1_lblImpVal::text').get()
            P = response.css('#ValueSummary1_pnlValue_lblLandVal::text').get()
            Q = response.css('#ValueSummary1_pnlValue_lblTotalVal::text').get()

            sheet["O" + repr(counter)] = response.css('#ValueSummary1_lblImpVal::text').get()
            sheet["P" + repr(counter)] = response.css('#ValueSummary1_pnlValue_lblLandVal::text').get()
            sheet["Q" + repr(counter)] = response.css('#ValueSummary1_pnlValue_lblTotalVal::text').get()



            #     Main Improvement (Current 2021)
            Improvement = response.xpath('//span[@id="MainImpRes1_lblMsg"]/p/text()').get()

            if Improvement == "No Main Improvement.":
                print("Nothing")
            else:
                AX = response.css('#MainImpRes1_lblBuildClass::text').get()
                BA = response.css('#MainImpRes1_lblConstrType::text').get()
                BS = response.css('#MainImpRes1_lblFullBath::text').get()
                BT = response.css('#MainImpRes1_lblHalfBath::text').get()
                CC = response.css('#MainImpRes1_lblYearBuilt::text').get()
                BB = response.css('#MainImpRes1_lblFoundType::text').get()
                BQ = response.css('#MainImpRes1_lblKitchen::text').get()
                AW = response.css('#MainImpRes1_lblCDU::text').get()
                BC = response.css('#MainImpRes1_lblRoofType::text').get()
                BR = response.css('#MainImpRes1_lblBedRoom::text').get()

                CB = response.css('#MainImpRes1_lblLivingArea::text').get()

                BD = response.css('#MainImpRes1_lblRoofMat::text').get()
                BJ = response.css('#MainImpRes1_lblWetBar::text').get()

                CA = response.css('#MainImpRes1_lblTotalArea::text').get()

                BE = response.css('#MainImpRes1_lblFence::text').get()
                BK = response.css('#MainImpRes1_lblFP::text').get()

                AV = response.css('#MainImpRes1_lblNumStories::text').get()

                BF = response.css('#MainImpRes1_lblExtWall::text').get()
                BL = response.css('#MainImpRes1_lblSprinkler::text').get()

                BM = response.css('#MainImpRes1_lblDeck::text').get()

                BH = response.css('#MainImpRes1_lblHeatType::text').get()
                BI = response.css('#MainImpRes1_lblAC::text').get()

                BN = response.css('#MainImpRes1_lblSpa::text').get()
                BO = response.css('#MainImpRes1_lblPool::text').get()
                BP = response.css('#MainImpRes1_lblSauna::text').get()

            # Additional Improvements (Current 2021)
            Additional_Improvement = response.xpath('//span[@id="ResImp1_lblMessage"]/p/text()').get()
            if Additional_Improvement == "No Additional Improvements.":
                print("Nothing")
                BW = None
                BX = None
                BZ = None
            else:
                BW = response.xpath('//table[@id="ResImp1_dgImp"]/tr[2]/td[2]/text()').get()
                BX = response.xpath('//table[@id="ResImp1_dgImp"]/tr[2]/td[3]/text()').get()
                BZ = response.xpath('//table[@id="ResImp1_dgImp"]/tr[2]/td[6]/text()').get()

            # Land (2021 Proposed Values)

            I = response.xpath('//table[@id="Land1_dgLand"]/tr[2]/td[2]/text()').get()
            H = response.xpath('//table[@id="Land1_dgLand"]/tr[2]/td[3]/text()').get()
            CE = response.xpath('//table[@id="Land1_dgLand"]/tr[2]/td[4]/text()').get()
            CF = response.xpath('//table[@id="Land1_dgLand"]/tr[2]/td[5]/text()').get()

            data = response.css("#Land1_dgLand_Label1_0::text").get().split()
            sq_acres = data[1]+' '+data[2]
            if sq_acres == 'SQUARE FEET':
                CP = response.css("#Land1_dgLand_Label1_0::text").get().split()[0]
                print("Value of Sq or Ac ",CP)
            else:
                CO = response.css("#Land1_dgLand_Label1_0::text").get().split()[0]
                print("Value of Sq or Ac ", CO)

            print("THis is value of ", AX)
            print("THis is value of ", BA)
            print("THis is value of ", BS)
            print("THis is value of ", BH)
            # print("THis is value of ", BI)
            # print("THis is value of ", BN)
            # print("THis is value of ", BO)
            # print("THis is value of ", BP)
            # print("THis is value of ", BT)
            # print("THis is value of ", BB)
            # print("THis is value of ", BQ)
            # print("THis is value of ", CC)
            # print("THis is value of ", AW)
            # print("THis is value of ", BC)
            # print("THis is value of ", BR)
            # print("THis is value of ", BJ)
            # print("THis is value of ", CB)
            # print("THis is value of ", BD)
            # print("THis is value of ", CA)
            # print("THis is value of ", BE)
            # print("THis is value of ", AV)
            # print("THis is value of ", BF)
            # print("THis is value of ", BL)
            # print("THis is value of ", BM)
            # print("THis is value of ", BK)
            #
            # print("Value of ", O)
            # print("Value of ", P)
            # print("Value of ", Q)
            # print("Value of ", AZ)
            #
            # print("Value of ", KS)
            # print("Value of ", KR)
            # print("Value of IA is",IA)
            #
            # print("Value of ", I)
            # print("Value of ", H)
            # print("Value of ", CE)
            # print("Value of ", CF)
            #
            # print("Value of BW", BW)
            # print("Value of BX", BX)
            # print("Value of BZ", BZ)

            counter += 1
            print("Value of Counter for this time is ",counter)

            # Save the spreadsheet
            workbook.save(filename="/Users/ARK/Desktop/output1.xlsx")
    def parse2(self, response):


        value = response.css("h3::text").getall()[8]
        if value == ' ':
            ID = response.css("h3::text").getall()[8].split(',')[0]
            IG = response.css("h3::text").getall()[8].split(',')[1]
        else:
            ID = response.css("h3::text").getall()[8].split(',')[0]
            IG = response.css("h3::text").getall()[8].split(',')[1]


        S = response.css('h3::text').getall()[37].strip()
        R = response.css('h3::text').getall()[39].strip()

        count = len(response.css('h3::text').getall()[41].strip().split())
        if count==2:
            HK = response.css('h3::text').getall()[41].strip().split()[0]
            HL = response.css('h3::text').getall()[41].strip().split()[1]
            HM = response.css('h3::text').getall()[41].strip().split()[2]
        elif count==3:
            HK = response.css('h3::text').getall()[41].strip().split()[0]
            HL = response.css('h3::text').getall()[41].strip().split()[1]

            HM = response.css('h3::text').getall()[41].strip().split()[2]
        else:
            HK = response.css('h3::text').getall()[41].strip().split()[0]



    def parse3(self, response):
        NA = response.css('h3::text').getall()[4].strip()

    def parse4(self, response):
        No_tax = response.css('h4::text').get()[:2]
        if No_tax == "No":
            print("Noting in Tax")
        else:
            MK = response.css('h3::text').getall()[16].strip()
            ML = response.css('h3::text').getall()[32].strip()
            MN = response.css('h3::text').getall()[48].strip()
            MO = response.css('h3::text').getall()[64].strip()
            MP = response.css('h3::text').getall()[80].strip()
            MQ = response.css('h3::text').getall()[96].strip()
            MS = response.css('h3>b::text').getall()[10].strip()








          # Start by opening the spreadsheet and selecting the main sheet
            # workbook = load_workbook(filename="output_file.xlsx")
            # sheet = workbook.active
            #
            # # Write what you want into a specific cell
            # sheet["C1"] = "writing"
            #
            # # Save the spreadsheet
            # workbook.save(filename="output_file.xlsx")

            # # Program extracting all columns
            # # name in Python
            # web = 'https://www.dallascad.org/AcctDetailRes.aspx?ID='
            #
            # loc = '/Users/ARK/Desktop/test1.xlsx'
            #
            # wb = xlrd.open_workbook(loc)
            # sheet = wb.sheet_by_index(0)
            # sheet.cell_value(0, 0)
            #
            # for i in range(sheet.nrows):
            #     web_name = web+sheet.cell_value(i, 0)
            #     print(web_name)
            #     print("Value of counter is ",i)
            #     if web_name is not None:
            #         yield response.follow(web_name, callback=self.parse)













